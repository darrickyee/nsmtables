# %%
import openpyxl as xl
import pandas as pd
import numpy as np

# %%

DATAPATH = 'U:/ESSIN Task 14/Mapping Report/2019/Standard Method/02_Mapping Tool Design/Data Source/SourceTable_States/SAS/Updated data/Long Files/Updated 0208'
OUTPATH = 'output'
dta = dict()
for gr in 'G4', 'G8':
    dta[gr] = pd.read_excel(f'{DATAPATH}/{gr}_long.xlsx')
    for col in 'nse', 'nse_se', 'nse_re':
        dta[gr][col] = pd.to_numeric(dta[gr][col], errors='coerce')

df = dta['G4'].append(dta['G8'])
del dta
# %%


def display_row(nse: float = np.nan,
                nse_se: float = np.nan,
                nse_re: float = np.nan,
                mark=np.nan):
    nse_nan = np.isnan(nse)
    return tuple((
        '\u2013' if nse_nan else nse,
        '†' if nse_nan else (
            '\u2013' if np.isnan(nse_se) else nse_se),
        '†' if nse_nan else (
            '\u2013' if np.isnan(nse_re) else nse_re),
        '!' if mark == '!' else ''
    ))


# %%
COL_OFFSETS = {
    'R': 3,
    'M': 7
}

STATES = tuple(dict.fromkeys(df['state']))


for state in STATES:
    wb = xl.load_workbook('source.xlsx')

    for grade in '4', '8':
        ws = wb['G'+grade]
        ws['B5'].value = ws['B5'].value.replace('[STATE_NAME]', state)

        data = df[(df['subjgrade'].str.endswith(grade)) &
                  (df['state'] == state)].sort_values('year')

        # Populate years (needs error checking?)
        for i, year in enumerate(dict.fromkeys(data['year'])):
            ws.cell(i+9, 2).value = year

        # Delete Not Applicable and relative error notes, if needed
        ws.row_dimensions[16].hidden = not any(
            np.isnan(x) for x in data['nse_se'].append(data['nse_re']))
        ws.row_dimensions[17].hidden = not any(
            np.isnan(x) for x in data['nse'])
        ws.row_dimensions[18].hidden = not any(x == '!' for x in data['mark'])

        for subj in 'R', 'M':

            j = COL_OFFSETS[subj]

            for i, row in enumerate(data[data['subjgrade'] == subj+grade].loc[:, 'nse':'mark'].itertuples()):
                # Start at row[1:] since "row" includes index
                outrow = display_row(*row[1:])
                for col in range(len(outrow)):
                    cell = ws.cell(i+9, col+j)
                    cell.value = outrow[col]

        wb.save(f'{OUTPATH}/{state}.xlsx')

# %%
